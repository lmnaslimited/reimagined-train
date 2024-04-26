from office365.runtime.auth.authentication_context import AuthenticationContext
from office365.sharepoint.client_context import ClientContext
from office365.sharepoint.files.file import File
import datetime
import openpyxl
import pandas as pd
import io
import copy
import warnings
import os
import time
import argparse
import json
import logging

warnings.filterwarnings('ignore')

shift = 10


def parse_arguments():
    parser = argparse.ArgumentParser(description='Script with -config and -encrypt arguments')

    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument('-config', metavar='CONFIG_ARG', nargs=1, help='Specify a single argument for configuration')
    group.add_argument('-encrypt', metavar='ENCRYPT_ARG', nargs=1, help='Specify a single argument for encryption')

    args = parser.parse_args()
    return args

def encrypt(message, shift):
    encrypted_message = ""
    for char in message:
        if char.isalpha():
            # Shift only alphabetical characters
            shift_amount = (ord(char) - ord('a') + shift) % 26 if char.islower() else (ord(char) - ord(
                'A') + shift) % 26
            encrypted_char = chr(ord('a') + shift_amount) if char.islower() else chr(ord('A') + shift_amount)
            encrypted_message += encrypted_char
        else:
            # Keep non-alphabetic characters unchanged
            encrypted_message += char
    return encrypted_message

def decrypt(encrypted_message, shift):
    return encrypt(encrypted_message, -shift)

class KpiWorkflowServer:

    def __init__(self, url_shrpt, username_shrpt, password_shrpt, input_file_path, output_file_name,
                 target_url, excel_buffer):
        self.url_shrpt = url_shrpt
        self.username_shrpt = username_shrpt
        self.password_shrpt = password_shrpt
        self.input_file_path = input_file_path
        self.target_url = target_url
        self.excel_buffer_path = excel_buffer
        self.excel_buffer_file_name = output_file_name
        self.excel_buffer = None
        self.ctx_auth = None
        self.file_size = -1
        self.previous_input_file_time_last_modified = None
        self.size_chunk = 1000000
        self.delay_seconds = 5
        self.authenticate()

    def authenticate(self):
        self.ctx_auth = AuthenticationContext(self.url_shrpt)
        if self.ctx_auth.acquire_token_for_user(self.username_shrpt, self.password_shrpt):
            ctx = ClientContext(self.url_shrpt, self.ctx_auth)
            web = ctx.web
            ctx.load(web)
            ctx.execute_query()
            logging.info('Authenticated into sharepoint as: ' + web.properties['Title'])
        else:
            logging.error(str(self.ctx_auth.get_last_error()))

    def upload_to_target_folder(self):
        ctx = ClientContext(self.url_shrpt, self.ctx_auth)
        target_folder = ctx.web.get_folder_by_server_relative_url(self.target_url)
        self.file_size = os.path.getsize(self.excel_buffer)
        try:
            uploaded_file = target_folder.files.create_upload_session(self.excel_buffer, self.size_chunk,
                                                                      self.print_upload_progress).execute_query()
            logging.info('File {0} has been uploaded successfully'.format(uploaded_file.serverRelativeUrl))
        except Exception as e:
            logging.error("Error while uploading to SharePoint:\n " + str(e))

    def print_upload_progress(self, offset):
        logging.info("Uploaded '{0}' bytes from '{1}'...[{2}%]".format(offset, self.file_size,
                                                                       round(offset / self.file_size * 100, 2)))

    def generate_excel_file_name(self):
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        folder_path = self.excel_buffer_path + "\\" + timestamp + "\\"
        while True:
            if os.path.exists(folder_path):
                timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
                folder_path = self.excel_buffer_path + "\\" + timestamp + "\\"
            else:
                break
        os.makedirs(folder_path)
        logging.info(f"Folder '{folder_path}' created successfully.")
        return folder_path + self.excel_buffer_file_name

    def write_to_excel(self):
        ctx = ClientContext(self.url_shrpt, self.ctx_auth)
        response = File.open_binary(ctx, self.input_file_path)
        logging.info("INFO: Extracting")
        list_of_sub_df = []
        list_of_sheet_names = []
        k = 10

        OI_based_on_market_seg_act_plan_map = \
            {"Building power infrastructure": {
                "Building power infrastructure (Actual MTD)": "Building power infrastructure (Plan MTD)"},
                "Renewables wind": {"Renewables wind (Actual MTD)": "Renewables wind (Plan MTD)"},
                "Renewables others": {"Renewables others  (Actual MTD)": "Renewables others  (Plan MTD)"},
                "Special W/O data centre": {
                    "Special W/O data centre (Actual MTD)": "Special W/O data centre (Plan MTD)"},
                "Data centre": {"Data centre (Actual MTD)": "Data centre (Plan MTD)"},
                "Lead sharing": {"Lead sharing (Actual MTD)": "Lead sharing (Plan MTD)"},
                "After sales": {"After sales  (Actual MTD)": "After sales  (Plan MTD)"},
                "400+ Project": {"400+ Project (Actual MTD)": "400+ Project (Plan MTD)"},
                "Middle east markets": {"Middle east markets (Actual MTD)": "Middle east markets (Plan MTD)"},
                "Intercompany ( IN/CN/US)": {
                    "Intercompany ( IN/CN/US) (Actual MTD)": "Intercompany ( IN/CN/US) (Plan MTD)"},
                "settled market (D,CZ,A,B,DK)": {
                    "settled market (D,CZ,A,B,DK) (Actual MTD)": "settled market (D,CZ,A,B,DK) (Plan MTD)"},
                "Near future growth ( UK, FR )": {
                    "Near future growth ( UK, FR ) (Actual MTD)": "Near future growth ( UK, FR ) (Plan MTD)"},
                "Price driven markets ( PL,H)": {
                    "Price driven markets ( PL,H) (Actual MTD)": "Price driven markets ( PL,H) (Plan MTD)"},
                "New markets(NL,E,S)": {"New markets(NL,E,S) (Actual MTD)": "New markets(NL,E,S) (Plan MTD)"},
                "Others": {"Others (Actual MTD)": "Others (Plan MTD)"},
                "Wind Baseline": {"Wind Baseline (Actual MTD)": "Wind Baseline (Plan MTD)"},
                "Metro": {"Metro (Actual MTD)": "Metro (Plan MTD)"},
                "LV casted": {"LV casted (Actual MTD)": "LV casted (Plan MTD)"},
                "Industry W/O data centre": {
                    "Industry W/O data centre (Actual MTD)": "Industry W/O data centre (Plan MTD)"},
                "Service": {"Service (Actual MTD)": "Service (Plan MTD)"},
                "CR US": {"CR US (Actual MTD)": "CR US (Plan MTD)"},
                "CR Intercompany": {"CR Intercompany (Actual MTD)": "CR Intercompany (Plan MTD)"},
                "Industry Special W/o DC": {
                    "Industry Special W/o DC (Actual MTD)": "Industry Special W/o DC (Plan MTD)"}}

        UnitMappingDictionary = {'SGB_CN_Input': 'SGB CN', 'SGBCR_Input': 'SGB CR', 'SGBMY_Input': 'SGB MY',
                                 'SGBCZ_input': "SGB CZ", 'SGBIN_Input ': "SGB IN", 'SGBUSA_input': "SGB USA",
                                 'BCV_input': "BCV"}

        logging.info("INFO: Opening the workbook")
        try:
            workbook = openpyxl.load_workbook(io.BytesIO(response.content), data_only=True)
            sheet_names = workbook.sheetnames
            kpi_data = None
            isFirstDfUpdated = False
            logging.info("INFO: Extracting")
            for sheet_name in sheet_names:
                sheet = workbook[sheet_name]
                date = []
                data = {}

                OI = ["OI (Actual MTD)", "OI (ActualYTD)", "OI (Plan MTD)", "OI (PlanYTD)", "OI (Actual vs Plan)",
                      "OI (Actual vs Plan YTD)"]

                sale = ["Sales (Actual)", "Sales (ActualYTD)", "Sales (Plan)", "Sales (PlanYTD)",
                        "Sales (Actual vs Plan)", "Sales (Actual vs Plan YTD)"]

                Sale_MVA = ["Sales MVA (Actual)", "Sales MVA (ActualYTD)", "Sales MVA (Plan)", "Sales MVA (PlanYTD)",
                            "Sales MVA (Actual vs Plan)", "Sales MVA (Actual vs Plan YTD)"]

                Sales_Units = ["Sales Units (Actual)", "Sales Units (ActualYTD)", "Sales Units (Plan)",
                               "Sales Units (PlanYTD)", "Sales Units (Actual vs Plan)",
                               "Sales Units (Actual vs Plan YTD)"]

                Sales_EUR_KvA = ["Sales €/kVA", "Sales €/kVA"]
                Sales_EUR_KvA_Count = 1

                Sales_MVA_Unit = ["Sales MVA/Unit", "Sales MVA/Unit"]
                Sales_MVA_Unit_Count = 1

                sale_ebitda = ["Sales EBITDA (Actual MTD)", "Sales EBITDA (Actual YTD)", "Sales EBITDA (Plan MTD)",
                               "Sales EBITDA (Plan YTD)", "Sales EBITDA (Actual vs Plan MTD)",
                               "Sales EBITDA (Actual vs Plan YTD)"]

                Material_ratio_total = ["Material ratio total (Actual MTD)", "Material ratio total (Actual YTD)",
                                        "Material ratio total (Plan MTD)", "Material ratio total (Plan YTD)"]

                TWC_value_I = ["TWC (Actual MTD)", "TWC (Plan MTD)", "TWC in % of Sales (Actual MTD)",
                               "TWC in % of Sales (Plan MTD)"]



                Production_MVA_non_PoC = ["Production MVA (Actual)", "Production MVA (ActualYTD)",
                                          "Production MVA (Plan)", "Production MVA (PlanYTD)"]

                Production_Units_non_PoC = ["Production Units (Actual)", "Production Units (ActualYTD)",
                                            "Production Units (Plan)", "Production Units (PlanYTD)"]

                Production_MVA_Unit_non_PoC = []

                Production_MVA_PoC = []

                Production_Units_PoC = []

                Production_MVA_Unit_PoC = []

                Production_hours = []

                Engineering_hours = []

                Total_hours = []

                Total_hours_per_MVA_produced_non_PoC = []

                Total_hours_per_Unit_produced_non_PoC = []

                Total_hours_per_MVA_produced_PoC = []

                Total_hours_per_Unit_produced_PoC = []

                Through_put_time_engineering = []

                Through_put_time_factory = []

                On_time_delivery_factory = []

                First_time_right_FTR = ["Number of failed transformer (Actual MTD)",
                                        "Number of failed transformer (Actual YTD)",
                                        "Number of tested transformer (Actual MTD)",
                                        "Number of tested transformer (Actual YTD)",
                                        "First time right (Actual MTD)",
                                        "First time right (Actual YTD)",
                                        "First time right (Plan MTD)"]

                First_pass_yield_FPY = ["Number of failed transformer (Actual MTD)",
                                        "Number of failed transformer (Actual YTD)",
                                        "Number of tested transformer (Actual MTD)",
                                        "Number of tested transformer (Actual YTD)",
                                        "First pass yield (Actual MTD)",
                                        "First pass yield (Actual YTD)",
                                        "First pass yield (Plan MTD)"]

                Failure_costs = ["Failure costs external (Actual MTD)", "Failure costs external (Actual YTD)",
                                 "Failure costs internal (Actual MTD)", "Failure costs internal (Actual YTD)",
                                 "Failure costs external + internal (Actual MTD)",
                                 "Failure costs external + internal (Actual YTD)",
                                 "Refunds failure costs external (Actual MTD)",
                                 "Refunds failure costs external (Actual YTD)",
                                 "Refunds failure costs internal (Actual MTD)",
                                 "Refunds failure costs internal (Actual YTD)",
                                 "Refunfs failure costs external + internal (Actual MTD)",
                                 "Refunfs failure costs external + internal (Actual YTD)",
                                 "Failure costs external + internal (Plan MTD)",
                                 "Failure costs external + internal (Plan YTD)"]

                OI_based_on_market_seg = []

                for value in OI_based_on_market_seg_act_plan_map.values():
                    for key, val in value.items():
                        OI_based_on_market_seg.append(val)
                        OI_based_on_market_seg.append(key)

                Order_back_logs = ["Order backlog value (Actual MTD)", "Order backlog value (Plan MTD)"]

                sheet_specific_market_seg = []

                current_column = ""
                is_getting_value = False

                # Access and print cell values
                index = 0
                head = ""

                for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
                    for cell in row:
                        if head == 'Month':
                            if index == 1:
                                index += 1
                                continue
                            else:
                                if type(cell.value) == datetime.datetime:
                                    date.append(cell.value)
                                else:
                                    index = 0
                                    head = ''
                                    data["Date"] = date

                        if head == 'Order Intake':
                            offset = 10
                            if (not is_getting_value) and (cell.value in OI):
                                current_column = cell.value
                                data[current_column] = []
                                OI.remove(cell.value)
                                is_getting_value = True

                            if is_getting_value and index > offset:
                                if len(data[current_column]) != len(data["Date"]):
                                    data[current_column].append(cell.value)
                                else:
                                    index = 1
                                    is_getting_value = False
                                    if len(OI) == 0:
                                        head = ''
                            elif is_getting_value:
                                index += 1

                        if head == 'OI market segment split':
                            offset = 10
                            if (not is_getting_value) and (cell.value in OI_based_on_market_seg):
                                current_column = cell.value
                                if current_column not in sheet_specific_market_seg:
                                    sheet_specific_market_seg.append(current_column)
                                data[current_column] = []
                                OI_based_on_market_seg.remove(cell.value)
                                is_getting_value = True

                            if is_getting_value and index > offset:
                                if len(data[current_column]) != len(data["Date"]):
                                    data[current_column].append(cell.value)
                                else:
                                    index = 1
                                    is_getting_value = False
                                    if len(OI_based_on_market_seg) == 0:
                                        head = ''
                            elif is_getting_value:
                                index += 1

                        if head == 'Sales EBITDA':
                            offset = 10
                            if (not is_getting_value) and (cell.value in sale_ebitda):
                                current_column = cell.value
                                data[current_column] = []
                                sale_ebitda.remove(cell.value)
                                is_getting_value = True

                            if is_getting_value and index > offset:
                                if len(data[current_column]) != len(data["Date"]):
                                    data[current_column].append(cell.value)
                                else:
                                    index = 1
                                    is_getting_value = False
                                    if len(sale_ebitda) == 0:
                                        head = ''
                            elif is_getting_value:
                                index += 1

                        if head == 'Sales':
                            offset = 10
                            if (not is_getting_value) and (cell.value in sale):
                                current_column = cell.value
                                data[current_column] = []
                                sale.remove(cell.value)
                                is_getting_value = True

                            if is_getting_value and index > offset:
                                if len(data[current_column]) != len(data["Date"]):
                                    data[current_column].append(cell.value)
                                else:
                                    index = 1
                                    is_getting_value = False
                                    if len(sale) == 0:
                                        head = ''
                            elif is_getting_value:
                                index += 1

                        if head == 'Order backlog':
                            offset = 10
                            if (not is_getting_value) and (cell.value in Order_back_logs):
                                current_column = cell.value
                                data[current_column] = []
                                Order_back_logs.remove(cell.value)
                                is_getting_value = True

                            if is_getting_value and index > offset:
                                if len(data[current_column]) != len(data["Date"]):
                                    data[current_column].append(cell.value)
                                else:
                                    index = 1
                                    is_getting_value = False
                                    if len(Order_back_logs) == 0:
                                        head = ''
                            elif is_getting_value:
                                index += 1

                        if head == 'Production MVA non PoC':
                            offset = 10
                            if (not is_getting_value) and (cell.value in Production_MVA_non_PoC):
                                current_column = cell.value
                                data[current_column] = []
                                Production_MVA_non_PoC.remove(cell.value)
                                is_getting_value = True

                            if is_getting_value and index > offset:
                                if len(data[current_column]) != len(data["Date"]):
                                    data[current_column].append(cell.value)
                                else:
                                    index = 1
                                    is_getting_value = False
                                    if len(Production_MVA_non_PoC) == 0:
                                        head = ''
                            elif is_getting_value:
                                index += 1

                        if head == 'Production Units non PoC':
                            offset = 10
                            if (not is_getting_value) and (cell.value in Production_Units_non_PoC):
                                current_column = cell.value
                                data[current_column] = []
                                Production_Units_non_PoC.remove(cell.value)
                                is_getting_value = True

                            if is_getting_value and index > offset:
                                if len(data[current_column]) != len(data["Date"]):
                                    data[current_column].append(cell.value)
                                else:
                                    index = 1
                                    is_getting_value = False
                                    if len(Production_Units_non_PoC) == 0:
                                        head = ''
                            elif is_getting_value:
                                index += 1

                        if head == 'Production MVA/Unit non PoC':
                            offset = 10
                            if (not is_getting_value) and (cell.value in Production_MVA_Unit_non_PoC):
                                current_column = cell.value
                                data[current_column] = []
                                Production_MVA_Unit_non_PoC.remove(cell.value)
                                is_getting_value = True

                            if is_getting_value and index > offset:
                                if len(data[current_column]) != len(data["Date"]):
                                    data[current_column].append(cell.value)
                                else:
                                    index = 1
                                    is_getting_value = False
                                    if len(Production_MVA_Unit_non_PoC) == 0:
                                        head = ''
                            elif is_getting_value:
                                index += 1

                        if head == 'Production MVA PoC':
                            offset = 10
                            if (not is_getting_value) and (cell.value in Production_MVA_PoC):
                                current_column = cell.value
                                data[current_column] = []
                                Production_MVA_PoC.remove(cell.value)
                                is_getting_value = True

                            if is_getting_value and index > offset:
                                if len(data[current_column]) != len(data["Date"]):
                                    data[current_column].append(cell.value)
                                else:
                                    index = 1
                                    is_getting_value = False
                                    if len(Production_MVA_PoC) == 0:
                                        head = ''
                            elif is_getting_value:
                                index += 1

                        if head == 'Production Units PoC':
                            offset = 10
                            if (not is_getting_value) and (cell.value in Production_Units_PoC):
                                current_column = cell.value
                                data[current_column] = []
                                Production_Units_PoC.remove(cell.value)
                                is_getting_value = True

                            if is_getting_value and index > offset:
                                if len(data[current_column]) != len(data["Date"]):
                                    data[current_column].append(cell.value)
                                else:
                                    index = 1
                                    is_getting_value = False
                                    if len(Production_Units_PoC) == 0:
                                        head = ''
                            elif is_getting_value:
                                index += 1

                        if head == 'Production MVA/Unit PoC':
                            offset = 10
                            if (not is_getting_value) and (cell.value in Production_MVA_Unit_PoC):
                                current_column = cell.value
                                data[current_column] = []
                                Production_MVA_Unit_PoC.remove(cell.value)
                                is_getting_value = True

                            if is_getting_value and index > offset:
                                if len(data[current_column]) != len(data["Date"]):
                                    data[current_column].append(cell.value)
                                else:
                                    index = 1
                                    is_getting_value = False
                                    if len(Production_MVA_Unit_PoC) == 0:
                                        head = ''
                            elif is_getting_value:
                                index += 1

                        if head == 'Production hours':
                            offset = 10
                            if (not is_getting_value) and (cell.value in Production_hours):
                                current_column = cell.value
                                data[current_column] = []
                                Production_hours.remove(cell.value)
                                is_getting_value = True

                            if is_getting_value and index > offset:
                                if len(data[current_column]) != len(data["Date"]):
                                    data[current_column].append(cell.value)
                                else:
                                    index = 1
                                    is_getting_value = False
                                    if len(Production_hours) == 0:
                                        head = ''
                            elif is_getting_value:
                                index += 1

                        if head == 'Engineering hours':
                            offset = 10
                            if (not is_getting_value) and (cell.value in Engineering_hours):
                                current_column = cell.value
                                data[current_column] = []
                                Engineering_hours.remove(cell.value)
                                is_getting_value = True

                            if is_getting_value and index > offset:
                                if len(data[current_column]) != len(data["Date"]):
                                    data[current_column].append(cell.value)
                                else:
                                    index = 1
                                    is_getting_value = False
                                    if len(Engineering_hours) == 0:
                                        head = ''
                            elif is_getting_value:
                                index += 1

                        if head == 'Total hours':
                            offset = 10
                            if (not is_getting_value) and (cell.value in Total_hours):
                                current_column = cell.value
                                data[current_column] = []
                                Total_hours.remove(cell.value)
                                is_getting_value = True

                            if is_getting_value and index > offset:
                                if len(data[current_column]) != len(data["Date"]):
                                    data[current_column].append(cell.value)
                                else:
                                    index = 1
                                    is_getting_value = False
                                    if len(Total_hours) == 0:
                                        head = ''
                            elif is_getting_value:
                                index += 1

                        if head == 'Total hours per MVA produced non PoC':
                            offset = 10
                            if (not is_getting_value) and (cell.value in Total_hours_per_MVA_produced_non_PoC):
                                current_column = cell.value
                                data[current_column] = []
                                Total_hours_per_MVA_produced_non_PoC.remove(cell.value)
                                is_getting_value = True

                            if is_getting_value and index > offset:
                                if len(data[current_column]) != len(data["Date"]):
                                    data[current_column].append(cell.value)
                                else:
                                    index = 1
                                    is_getting_value = False
                                    if len(Total_hours_per_MVA_produced_non_PoC) == 0:
                                        head = ''
                            elif is_getting_value:
                                index += 1

                        if head == 'Total hours per Unit produced non PoC':
                            offset = 10
                            if (not is_getting_value) and (cell.value in Total_hours_per_Unit_produced_non_PoC):
                                current_column = cell.value
                                data[current_column] = []
                                Total_hours_per_Unit_produced_non_PoC.remove(cell.value)
                                is_getting_value = True

                            if is_getting_value and index > offset:
                                if len(data[current_column]) != len(data["Date"]):
                                    data[current_column].append(cell.value)
                                else:
                                    index = 1
                                    is_getting_value = False
                                    if len(Total_hours_per_Unit_produced_non_PoC) == 0:
                                        head = ''
                            elif is_getting_value:
                                index += 1

                        if head == 'Total hours per MVA produced PoC':
                            offset = 10
                            if (not is_getting_value) and (cell.value in Total_hours_per_MVA_produced_PoC):
                                current_column = cell.value
                                data[current_column] = []
                                Total_hours_per_MVA_produced_PoC.remove(cell.value)
                                is_getting_value = True

                            if is_getting_value and index > offset:
                                if len(data[current_column]) != len(data["Date"]):
                                    data[current_column].append(cell.value)
                                else:
                                    index = 1
                                    is_getting_value = False
                                    if len(Total_hours_per_MVA_produced_PoC) == 0:
                                        head = ''
                            elif is_getting_value:
                                index += 1

                        if head == 'Total hours per Unit produced PoC':
                            offset = 10
                            if (not is_getting_value) and (cell.value in Total_hours_per_Unit_produced_PoC):
                                current_column = cell.value
                                data[current_column] = []
                                Total_hours_per_Unit_produced_PoC.remove(cell.value)
                                is_getting_value = True

                            if is_getting_value and index > offset:
                                if len(data[current_column]) != len(data["Date"]):
                                    data[current_column].append(cell.value)
                                else:
                                    index = 1
                                    is_getting_value = False
                                    if len(Total_hours_per_Unit_produced_PoC) == 0:
                                        head = ''
                            elif is_getting_value:
                                index += 1

                        if head == 'Through put time engineering':
                            offset = 10
                            if (not is_getting_value) and (cell.value in Through_put_time_engineering):
                                current_column = cell.value
                                data[current_column] = []
                                Through_put_time_engineering.remove(cell.value)
                                is_getting_value = True

                            if is_getting_value and index > offset:
                                if len(data[current_column]) != len(data["Date"]):
                                    data[current_column].append(cell.value)
                                else:
                                    index = 1
                                    is_getting_value = False
                                    if len(Through_put_time_engineering) == 0:
                                        head = ''
                            elif is_getting_value:
                                index += 1

                        if head == 'Through put time factory':
                            offset = 10
                            if (not is_getting_value) and (cell.value in Through_put_time_factory):
                                current_column = cell.value
                                data[current_column] = []
                                Through_put_time_factory.remove(cell.value)
                                is_getting_value = True

                            if is_getting_value and index > offset:
                                if len(data[current_column]) != len(data["Date"]):
                                    data[current_column].append(cell.value)
                                else:
                                    index = 1
                                    is_getting_value = False
                                    if len(Through_put_time_factory) == 0:
                                        head = ''
                            elif is_getting_value:
                                index += 1

                        if head == 'First time right (FTR)':
                            offset = 10
                            if (not is_getting_value) and (cell.value in First_time_right_FTR):
                                current_column = cell.value
                                data[current_column] = []
                                First_time_right_FTR.remove(cell.value)
                                is_getting_value = True

                            if is_getting_value and index > offset:
                                if len(data[current_column]) != len(data["Date"]):
                                    data[current_column].append(cell.value)
                                else:
                                    index = 1
                                    is_getting_value = False
                                    if len(First_time_right_FTR) == 0:
                                        head = ''
                            elif is_getting_value:
                                index += 1

                        if head == 'First pass yield (FPY)':
                            offset = 10
                            if (not is_getting_value) and (cell.value in First_pass_yield_FPY):
                                current_column = cell.value
                                data[current_column] = []
                                First_pass_yield_FPY.remove(cell.value)
                                is_getting_value = True

                            if is_getting_value and index > offset:
                                if len(data[current_column]) != len(data["Date"]):
                                    data[current_column].append(cell.value)
                                else:
                                    index = 1
                                    is_getting_value = False
                                    if len(First_pass_yield_FPY) == 0:
                                        head = ''
                            elif is_getting_value:
                                index += 1

                        if head == 'Failure costs':
                            offset = 10
                            if (not is_getting_value) and (cell.value in Failure_costs):
                                current_column = cell.value
                                data[current_column] = []
                                Failure_costs.remove(cell.value)
                                is_getting_value = True

                            if is_getting_value and index > offset:
                                if len(data[current_column]) != len(data["Date"]):
                                    data[current_column].append(cell.value)
                                else:
                                    index = 1
                                    is_getting_value = False
                                    if len(Failure_costs) == 0:
                                        head = ''
                            elif is_getting_value:
                                index += 1

                        if head == 'Sales MVA':
                            offset = 10
                            if (not is_getting_value) and (cell.value in Sale_MVA):
                                current_column = cell.value
                                data[current_column] = []
                                Sale_MVA.remove(cell.value)
                                is_getting_value = True

                            if is_getting_value and index > offset:
                                if len(data[current_column]) != len(data["Date"]):
                                    data[current_column].append(cell.value)
                                else:
                                    index = 1
                                    is_getting_value = False
                                    if len(Sale_MVA) == 0:
                                        head = ''
                            elif is_getting_value:
                                index += 1

                        if head == 'Sales Units':
                            offset = 10
                            if (not is_getting_value) and (cell.value in Sales_Units):
                                current_column = cell.value
                                data[current_column] = []
                                Sales_Units.remove(cell.value)
                                is_getting_value = True

                            if is_getting_value and index > offset:
                                if len(data[current_column]) != len(data["Date"]):
                                    data[current_column].append(cell.value)
                                else:
                                    index = 1
                                    is_getting_value = False
                                    if len(Sales_Units) == 0:
                                        head = ''
                            elif is_getting_value:
                                index += 1

                        if head == 'Sales EUR/KvA':
                            offset = 10
                            if (not is_getting_value) and (cell.value in Sales_EUR_KvA):
                                current_column = cell.value + str(Sales_EUR_KvA_Count)
                                Sales_EUR_KvA_Count +=1
                                data[current_column] = []
                                Sales_EUR_KvA.remove(cell.value)
                                is_getting_value = True

                            if is_getting_value and index > offset:
                                if len(data[current_column]) != len(data["Date"]):
                                    data[current_column].append(cell.value)
                                else:
                                    index = 1
                                    is_getting_value = False
                                    if len(Sales_EUR_KvA) == 0:
                                        head = ''
                            elif is_getting_value:
                                index += 1

                        if head == 'Sales MVA/Unit':
                            offset = 9
                            if (not is_getting_value) and (cell.value in Sales_MVA_Unit):
                                current_column = cell.value + str(Sales_MVA_Unit_Count)
                                Sales_MVA_Unit_Count += 1
                                data[current_column] = []
                                Sales_MVA_Unit.remove(cell.value)
                                is_getting_value = True

                            if is_getting_value and index > offset:
                                if len(data[current_column]) != len(data["Date"]):
                                    data[current_column].append(cell.value)
                                else:
                                    index = 1
                                    is_getting_value = False
                                    if len(Sales_MVA_Unit) == 0:
                                        head = ''
                            elif is_getting_value:
                                index += 1

                        if head == 'Material ratio total':
                            offset = 10
                            if (not is_getting_value) and (cell.value in Material_ratio_total):
                                current_column = cell.value
                                data[current_column] = []
                                Material_ratio_total.remove(cell.value)
                                is_getting_value = True

                            if is_getting_value and index > offset:
                                if len(data[current_column]) != len(data["Date"]):
                                    data[current_column].append(cell.value)
                                else:
                                    index = 1
                                    is_getting_value = False
                                    if len(Material_ratio_total) == 0:
                                        head = ''
                            elif is_getting_value:
                                index += 1

                        if head == 'TWC value (I)':
                            offset = 10
                            if (not is_getting_value) and (cell.value in TWC_value_I):
                                current_column = cell.value
                                data[current_column] = []
                                TWC_value_I.remove(cell.value)
                                is_getting_value = True

                            if is_getting_value and index > offset:
                                if len(data[current_column]) != len(data["Date"]):
                                    data[current_column].append(cell.value)
                                else:
                                    index = 1
                                    is_getting_value = False
                                    if len(TWC_value_I) == 0:
                                        head = ''
                            elif is_getting_value:
                                index += 1

                        if cell.value == 'Month':
                            head = cell.value
                            index = 1

                        if cell.value == 'Order Intake':
                            head = cell.value
                            index = 1

                        if cell.value == 'Sales':
                            head = cell.value
                            index = 1

                        if cell.value == 'Sales EBITDA':
                            head = cell.value
                            index = 1

                        if cell.value == 'OI market segment split':
                            head = cell.value
                            index = 1

                        if cell.value == 'Order backlog':
                            head = cell.value
                            index = 1

                        if cell.value == 'Production MVA non PoC':
                            head = cell.value
                            index = 1

                        if cell.value == 'Production Units non PoC':
                            head = cell.value
                            index = 1

                        if cell.value == 'Production MVA/Unit non PoC':
                            head = cell.value
                            index = 1

                        if cell.value == 'Production MVA PoC':
                            head = cell.value
                            index = 1

                        if cell.value == 'Production Units PoC':
                            head = cell.value
                            index = 1

                        if cell.value == 'Production MVA/Unit PoC':
                            head = cell.value
                            index = 1

                        if cell.value == 'Production hours':
                            head = cell.value
                            index = 1

                        if cell.value == 'Engineering hours':
                            head = cell.value
                            index = 1

                        if cell.value == 'Total hours':
                            head = cell.value
                            index = 1

                        if cell.value == 'Total hours per MVA produced non PoC':
                            head = cell.value
                            index = 1

                        if cell.value == 'Total hours per Unit produced non PoC':
                            head = cell.value
                            index = 1

                        if cell.value == 'Total hours per MVA produced PoC':
                            head = cell.value
                            index = 1

                        if cell.value == 'Total hours per Unit produced PoC':
                            head = cell.value
                            index = 1

                        if cell.value == 'Through put time engineering':
                            head = cell.value
                            index = 1

                        if cell.value == 'Through put time factory':
                            head = cell.value
                            index = 1

                        if cell.value == 'First time right (FTR)':
                            head = cell.value
                            index = 1

                        if cell.value == 'First pass yield (FPY)':
                            head = cell.value
                            index = 1

                        if cell.value == 'Failure costs':
                            head = cell.value
                            index = 1

                        if cell.value == 'Sales MVA':
                            head = cell.value
                            index = 1

                        if cell.value == 'Sales Units':
                            head = cell.value
                            index = 1

                        if cell.value == 'Sales EUR/KvA':
                            head = cell.value
                            index = 1

                        if cell.value == 'Sales MVA/Unit':
                            head = cell.value
                            index = 1

                        if cell.value == 'Material ratio total':
                            head = cell.value
                            index = 1

                        if cell.value == 'TWC value (I)':
                            head = cell.value
                            index = 1

                dataframe_temp = pd.DataFrame(data)

                if sheet_name in UnitMappingDictionary.keys():
                    dataframe_temp["Unit"] = UnitMappingDictionary[sheet_name]
                else:
                    continue
                list_of_sheet_names.append(sheet_name)
                sheet_specific_market_seg.insert(0, "Unit")
                sheet_specific_market_seg.insert(1, "Date")
                try:
                    list_of_sub_df.append(dataframe_temp[sheet_specific_market_seg])
                except:
                    None
                if not isFirstDfUpdated:
                    kpi_data = dataframe_temp
                    isFirstDfUpdated = True
                else:
                    kpi_data = pd.concat([kpi_data, dataframe_temp])

            logging.info("INFO: Extraction completed")

            # Close the workbook
            workbook.close()
            kpi_data = kpi_data[kpi_data['Date'].dt.year <= 2028]
            kpi_data.fillna(0, inplace=True)
            kpi_data["Year"] = kpi_data["Date"].dt.year
            kpi_data["Month"] = kpi_data["Date"].dt.strftime('%m-%b')

            logging.info("INFO: The file closed")

            transformed_dfs = []
            for df in list_of_sub_df:
                transform = pd.DataFrame(columns=["Unit", "Date", "Market Segment", "(Actual MTD)", "(Plan MTD)"])
                for index, row in df.iterrows():
                    for Market_segment, value in OI_based_on_market_seg_act_plan_map.items():
                        for actual, plan in value.items():
                            if (actual in df.columns) and (plan in df.columns):
                                transform.loc[len(transform)] = [row[transform.columns[0]], row[transform.columns[1]],
                                                                 Market_segment, row[actual], row[plan]]
                            elif actual in df.columns:
                                transform.loc[len(transform)] = [row[transform.columns[0]], row[transform.columns[1]],
                                                                 Market_segment, row[actual], ""]
                            elif plan in df.columns:
                                transform.loc[len(transform)] = [row[transform.columns[0]], row[transform.columns[1]],
                                                                 Market_segment, "", row[plan]]
                transformed_dfs.append(transform)

            transformed_dfs_with_additional_columns = []

            for df in transformed_dfs:
                df = df[df['Date'].dt.year <= 2028]
                df["Year"] = df["Date"].dt.year
                df["Month"] = df["Date"].dt.strftime('%m-%b')
                df["Rel"] = df["Unit"] + df["Year"].astype(str) + df["Month"]
                transformed_dfs_with_additional_columns.append(df)

            Market_segments = pd.DataFrame(columns=["Unit", "Date", "Market Segment"])

            for df in transformed_dfs_with_additional_columns:
                Market_segments = \
                    pd.concat(
                        [Market_segments, df[["Unit", "Date", "Market Segment", "(Actual MTD)", "(Plan MTD)", "Rel"]]])

            Market_segments.fillna(0, inplace=True)

            del dataframe_temp

            kpi_data.reset_index(inplace=True, drop=True)
            kpi_data["Year"] = kpi_data["Date"].dt.year
            kpi_data["Month"] = kpi_data["Date"].dt.strftime('%m-%b')
            units = kpi_data["Unit"].unique()
            years = kpi_data["Year"].unique()
            months = kpi_data["Month"].unique()
            AllRepeated = copy.deepcopy(kpi_data)

            for unit in units:
                for year in years:
                    new_row = {"OI (Actual MTD)": {"OI (ActualYTD)": 0}, "OI (Plan MTD)": {"OI (PlanYTD)": 0},
                               "Sales (Actual)": {"Sales (ActualYTD)": 0},
                               "Sales (Plan)": {"Sales (PlanYTD)": 0},
                               "Sales EBITDA (Actual MTD)": {"Sales EBITDA (Actual YTD)": 0},
                               "Sales EBITDA (Plan MTD)": {"Sales EBITDA (Plan YTD)": 0},
                               "Sales MVA (Actual)": {"Sales MVA (ActualYTD)": 0},
                               "Sales MVA (Plan)": {"Sales MVA (PlanYTD)": 0},
                               "Sales Units (Actual)": {"Sales Units (ActualYTD)": 0},
                               "Sales Units (Plan)": {"Sales Units (PlanYTD)": 0},
                               "Production MVA (Actual)" : {"Production MVA (ActualYTD)": 0},
                               "Production MVA (Plan)" : {"Production MVA (PlanYTD)": 0},
                               "Production Units (Actual)" : {"Production Units (ActualYTD)": 0},
                               "Production Units (Plan)" : {"Production Units (PlanYTD)": 0},
                               "Material ratio total (Actual MTD)": {"Material ratio total (Actual YTD)": 0},
                               "Material ratio total (Plan MTD)": {"Material ratio total (Plan YTD)": 0}
                               }

                    df = kpi_data.loc[(kpi_data["Unit"] == unit) & (kpi_data["Year"] == year)]
                    last_index = 0
                    last_row = None
                    for index, row in df.iterrows():
                        last_index = index
                        last_row = row
                        for value in new_row.values():
                            for key in value.keys():
                                if row[key] != 0:
                                    value[key] = row[key]
                    last_index += 1
                    if last_row is not None:
                        for key in new_row.keys():
                            for key2, value in new_row[key].items():
                                last_row[key] = value
                                last_row[key2] = value
                        last_row["Month"] = "All"
                        kpi_data.loc[len(kpi_data)] = last_row

            AllRepeated["Month filter"] = AllRepeated["Month"]
            kpi_data_all = copy.deepcopy(AllRepeated)
            kpi_data_all["Month filter"] = "All"
            kpi_data_all = pd.concat([kpi_data_all, AllRepeated])

            kpi_data_all["Rel"] = kpi_data_all["Unit"] + kpi_data_all["Year"].astype(str) + kpi_data_all["Month filter"]

            kpi_data["Rel"] = kpi_data["Unit"] + kpi_data["Year"].astype(str) + kpi_data["Month"]

            OI_based_on_market_seg = []

            for value in OI_based_on_market_seg_act_plan_map.values():
                for key, val in value.items():
                    OI_based_on_market_seg.append(val)
                    OI_based_on_market_seg.append(key)

            kpi_data.drop(columns=OI_based_on_market_seg, inplace=True)
            kpi_data_all.drop(columns=OI_based_on_market_seg, inplace=True)

            kpi_data.columns = ['Date', 'OI (Actual MTD)', 'OI (ActualYTD)', 'OI (Plan MTD)',
               'OI (PlanYTD)', 'OI (Actual vs Plan)', 'OI (Actual vs Plan YTD)',
               'Sales (Actual)', 'Sales (ActualYTD)', 'Sales (Plan)',
               'Sales (PlanYTD)', 'Sales (Actual vs Plan)',
               'Sales (Actual vs Plan YTD)', 'Sales MVA (Actual)',
               'Sales MVA (ActualYTD)', 'Sales MVA (Plan)', 'Sales MVA (PlanYTD)',
               'Sales MVA (Actual vs Plan)', 'Sales MVA (Actual vs Plan YTD)',
               'Sales Units (Actual)', 'Sales Units (ActualYTD)', 'Sales Units (Plan)',
               'Sales Units (PlanYTD)', 'Sales Units (Actual vs Plan)',
               'Sales Units (Actual vs Plan YTD)', 'Sales €/kVA (Actual)', 'Sales €/kVA (Plan)',
               'Sales MVA/Unit (Actual)', 'Sales MVA/Unit (Plan)', 'Sales EBITDA (Actual MTD)',
               'Sales EBITDA (Actual YTD)', 'Sales EBITDA (Plan MTD)',
               'Sales EBITDA (Plan YTD)', 'Sales EBITDA (Actual vs Plan MTD)',
               'Sales EBITDA (Actual vs Plan YTD)', 'Production MVA (Actual)',
               'Production MVA (ActualYTD)', 'Production MVA (Plan)',
               'Production MVA (PlanYTD)', 'Production Units (Actual)',
               'Production Units (ActualYTD)', 'Production Units (Plan)',
               'Production Units (PlanYTD)',
               'Number of failed transformer (Actual MTD)',
               'Number of failed transformer (Actual YTD)',
               'Number of tested transformer (Actual MTD)',
               'Number of tested transformer (Actual YTD)',
               'First time right (Actual MTD)', 'First time right (Actual YTD)',
               'First time right (Plan MTD)', 'First pass yield (Actual MTD)',
               'First pass yield (Actual YTD)', 'First pass yield (Plan MTD)',
               'Failure costs external (Actual MTD)',
               'Failure costs external (Actual YTD)',
               'Failure costs internal (Actual MTD)',
               'Failure costs internal (Actual YTD)',
               'Failure costs external + internal (Actual MTD)',
               'Failure costs external + internal (Actual YTD)',
               'Refunds failure costs external (Actual MTD)',
               'Refunds failure costs external (Actual YTD)',
               'Refunds failure costs internal (Actual MTD)',
               'Refunds failure costs internal (Actual YTD)',
               'Refunfs failure costs external + internal (Actual MTD)',
               'Refunfs failure costs external + internal (Actual YTD)',
               'Failure costs external + internal (Plan MTD)',
               'Failure costs external + internal (Plan YTD)',
               'Material ratio total (Actual MTD)',
               'Material ratio total (Actual YTD)', 'Material ratio total (Plan MTD)',
               'Material ratio total (Plan YTD)', 'TWC (Actual MTD)', 'TWC (Plan MTD)',
               'TWC in % of Sales (Actual MTD)', 'TWC in % of Sales (Plan MTD)',
               'Order backlog value (Actual MTD)', 'Order backlog value (Plan MTD)',
               'Unit', 'Year', 'Month', 'Rel']

            kpi_data_all.columns = ['Date', 'OI (Actual MTD)', 'OI (ActualYTD)', 'OI (Plan MTD)',
               'OI (PlanYTD)', 'OI (Actual vs Plan)', 'OI (Actual vs Plan YTD)',
               'Sales (Actual)', 'Sales (ActualYTD)', 'Sales (Plan)',
               'Sales (PlanYTD)', 'Sales (Actual vs Plan)',
               'Sales (Actual vs Plan YTD)', 'Sales MVA (Actual)',
               'Sales MVA (ActualYTD)', 'Sales MVA (Plan)', 'Sales MVA (PlanYTD)',
               'Sales MVA (Actual vs Plan)', 'Sales MVA (Actual vs Plan YTD)',
               'Sales Units (Actual)', 'Sales Units (ActualYTD)', 'Sales Units (Plan)',
               'Sales Units (PlanYTD)', 'Sales Units (Actual vs Plan)',
               'Sales Units (Actual vs Plan YTD)', 'Sales €/kVA (Actual)', 'Sales €/kVA (Plan)',
               'Sales MVA/Unit (Actual)', 'Sales MVA/Unit (Plan)', 'Sales EBITDA (Actual MTD)',
               'Sales EBITDA (Actual YTD)', 'Sales EBITDA (Plan MTD)',
               'Sales EBITDA (Plan YTD)', 'Sales EBITDA (Actual vs Plan MTD)',
               'Sales EBITDA (Actual vs Plan YTD)', 'Production MVA (Actual)',
               'Production MVA (ActualYTD)', 'Production MVA (Plan)',
               'Production MVA (PlanYTD)', 'Production Units (Actual)',
               'Production Units (ActualYTD)', 'Production Units (Plan)',
               'Production Units (PlanYTD)',
               'Number of failed transformer (Actual MTD)',
               'Number of failed transformer (Actual YTD)',
               'Number of tested transformer (Actual MTD)',
               'Number of tested transformer (Actual YTD)',
               'First time right (Actual MTD)', 'First time right (Actual YTD)',
               'First time right (Plan MTD)', 'First pass yield (Actual MTD)',
               'First pass yield (Actual YTD)', 'First pass yield (Plan MTD)',
               'Failure costs external (Actual MTD)',
               'Failure costs external (Actual YTD)',
               'Failure costs internal (Actual MTD)',
               'Failure costs internal (Actual YTD)',
               'Failure costs external + internal (Actual MTD)',
               'Failure costs external + internal (Actual YTD)',
               'Refunds failure costs external (Actual MTD)',
               'Refunds failure costs external (Actual YTD)',
               'Refunds failure costs internal (Actual MTD)',
               'Refunds failure costs internal (Actual YTD)',
               'Refunfs failure costs external + internal (Actual MTD)',
               'Refunfs failure costs external + internal (Actual YTD)',
               'Failure costs external + internal (Plan MTD)',
               'Failure costs external + internal (Plan YTD)',
               'Material ratio total (Actual MTD)',
               'Material ratio total (Actual YTD)', 'Material ratio total (Plan MTD)',
               'Material ratio total (Plan YTD)', 'TWC (Actual MTD)', 'TWC (Plan MTD)',
               'TWC in % of Sales (Actual MTD)', 'TWC in % of Sales (Plan MTD)',
               'Order backlog value (Actual MTD)', 'Order backlog value (Plan MTD)',
               'Unit', 'Year', 'Month', 'Month filter', 'Rel']

            # Write the DataFrames to the in-memory buffer as Excel sheets
            with pd.ExcelWriter(self.excel_buffer, engine='xlsxwriter') as writer:
                kpi_data.to_excel(writer, index=False, sheet_name='kpi_data', encoding='utf-8')
                kpi_data_all.to_excel(writer, index=False, sheet_name='kpi_data_all', encoding='utf-8')
                Market_segments.to_excel(writer, index=False, sheet_name='Market_segments', encoding='utf-8')
                writer.save()

            del transform
            del AllRepeated
            del df
        except Exception as e:
            logging.error("Exception raised: " + str(e))

    def is_file_updated(self):
        ctx = ClientContext(self.url_shrpt, self.ctx_auth)
        file = ctx.web.get_file_by_server_relative_path(self.input_file_path)
        ctx.load(file)
        ctx.execute_query()
        current_input_file_time_last_modified = file.time_last_modified

        if self.previous_input_file_time_last_modified:
            current_file_modified_date = \
                datetime.datetime.strptime(
                    current_input_file_time_last_modified.strftime("%Y-%m-%dT%H:%M:%SZ"),
                    "%Y-%m-%dT%H:%M:%SZ")
            previous_file_modified_date = \
                datetime.datetime.strptime(
                    self.previous_input_file_time_last_modified.strftime("%Y-%m-%dT%H:%M:%SZ"),
                    "%Y-%m-%dT%H:%M:%SZ")

            if current_file_modified_date <= previous_file_modified_date:
                time.sleep(self.delay_seconds)
                return False
            else:
                self.previous_input_file_time_last_modified = current_input_file_time_last_modified
        else:
            self.previous_input_file_time_last_modified = current_input_file_time_last_modified
        return True

    def run(self):
        logging.info("Server is watching the input file " + input_file_path)
        while True:
            try:
                if self.is_file_updated():
                    logging.info("The File update detected, triggering the workflow")
                    self.excel_buffer = self.generate_excel_file_name()
                    self.write_to_excel()
                    logging.info("Write successfull")
                    self.upload_to_target_folder()
                    logging.info("Server is watching the input file " + input_file_path)
            except KeyboardInterrupt as e:
                logging.info("\nCtrl+C detected. Stopping the server")
                break


if __name__ == "__main__":
    args = parse_arguments()
    if args.encrypt:
        print(encrypt(args.encrypt[0], shift))
    elif args.config:
        try:
            # Open the file in read mode
            with open(args.config[0], 'r') as file:
                # Use json.load() to read the object from the file
                loaded_object = json.load(file)

            url_shrpt = loaded_object["url_shrpt"]
            username_shrpt = loaded_object["username_shrpt"]
            password_shrpt = decrypt(loaded_object["password_shrpt"], shift)
            input_file_path = loaded_object["input_file_path"]
            output_file_name = loaded_object["output_file_name"]
            target_url = loaded_object["target_url"]
            excel_buffer = loaded_object["excel_buffer"]

            # Configure logging to write to a log file
            log_file_path = loaded_object["log_file_path"]
            logging.basicConfig(filename=log_file_path, level=logging.INFO,
                                format='%(asctime)s - %(levelname)s - %(message)s - %(lineno)d')

            # Create a console handler and set its level to INFO
            console_handler = logging.StreamHandler()
            console_handler.setLevel(logging.INFO)
            # Create a formatter and attach it to the console handler
            formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s - %(lineno)d')
            console_handler.setFormatter(formatter)

            # Add the console handler to the root logger
            logging.getLogger().addHandler(console_handler)

            kpi_workflow_server = KpiWorkflowServer(url_shrpt, username_shrpt, password_shrpt, input_file_path,
                                                    output_file_name, target_url, excel_buffer)
            kpi_workflow_server.run()

            # try:
            #     kpi_workflow_server = KpiWorkflowServer(url_shrpt, username_shrpt, password_shrpt, input_file_path,
            #                                             output_file_name, target_url, excel_buffer)
            #     kpi_workflow_server.run()
            # except Exception as e:
            #     logging.error(str(e))

        except Exception as e:
            print(e)
